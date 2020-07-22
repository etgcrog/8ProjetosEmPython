from selenium import webdriver
import openpyxl
from time import sleep
import pyautogui
import os
import shutil

class Extracao:
    def __init__(self):
        self.chrome_options = webdriver.ChromeOptions()
        arguments = ["--disable-notifications","--lang=pt-BR"]
        for argument in arguments:
            self.chrome_options.add_argument(argument)
        caminho_padrao_para_download = '/home/duduka/Área de Trabalho/Boleto'  # Coloque o caminho de onde seus downloads devem ir
        self.chrome_options.add_experimental_option("prefs", {
            'download.default_directory': caminho_padrao_para_download,
            # Atualiza diretório para dietório passado acima
            'download.directory_upgrade': True,
            # Seta se o navegar deve pedir ou não para fazer download
            'download.promp_for_download': False,
            "profile.default_content_setting_values.notifications": 2,  # Desabilita notificações
            # Allow multiple downoads
            "profile.default_content_setting_values.automatic_downloads": 1,  # Permite multiplos downloads automático
        })
        self.driver = webdriver.Chrome(executable_path='./chromedriver', options= self.chrome_options)
        self.driver.implicitly_wait(10)

    def criarplanilha(self):
        self.planilha = openpyxl.Workbook()
        self.planilha.create_sheet('extração')
        self.planilha_dados = self.planilha['extração']
        self.planilha_dados.cell(row = 1, column= 1, value = '')
        self.planilha_dados.cell(row = 1, column= 2, value = '')
        self.planilha_dados.cell(row = 1, column= 3, value = '')

    def scraping(self):
        try:
            while True:
                self.entar = self.driver.find_element_by_xpath('//*[@id="tab-1"]/h3')
                self.entar.click()
                self.login = self.driver.find_element_by_id('login')
                self.login.send_keys('#LOGIN')
                self.senha = self.driver.find_element_by_id('senha')
                self.senha.send_keys('#SENHA')
                self.btn = self.driver.find_element_by_id('bt_enviar_autenticacao_')
                self.btn.click()
                self.btn_continuar1 = self.driver.find_element_by_xpath('/html/body/div/div/a')
                self.btn_continuar1.click()
                self.btn_continuar2 = self.driver.find_element_by_xpath('//*[@id="conteudo"]/div/div/div/p/a')
                self.btn_continuar2.click()
                sleep(1)
                self.btn_continuar3 = self.driver.find_element_by_id('btn-prosseguir-pendencia')
                self.btn_continuar3.click()
                sleep(1)
                self.btn_fechar = self.driver.find_element_by_xpath('//*[@id="fecha_popup"]')
                self.btn_fechar.click()
                sleep(1)
                self.financeiro = self.driver.find_element_by_xpath('//*[@id="menu"]/ul/li[4]/a')
                self.financeiro.click()
                sleep(1)
                self.mensalidades = self.driver.find_element_by_xpath('//*[@id="menu"]/ul/li[4]/ul/li[3]/a')
                self.mensalidades.click()
                # self.table_mensalidade = self.driver.find_elements_by_xpath("//tr[@class='table2020_2']")
                self.valor_mensalidade = self.driver.find_elements_by_xpath("//td[@class='text-center']")

                self.salvar_boleto()
                # self.organizarmensalidades()
                # self.driver.quit()

                # self.criarplanilha()
                # self.mensalidade_do_mes()
                break

        except Exception as erro:
            print('Não há mais informações para extrair')
            print(erro)

    def salvar_boleto(self):
        pyautogui.click(x=314, y= 452, button='left', duration=2)
        sleep(3)
        pyautogui.click(x=831, y= 197, button='left', duration=2)
        pyautogui.click(x=830, y= 669, button='left', duration=2)
        # pyautogui.click(x=1062 ,y=704, button='left', duration=2)
        # pyautogui.alert(title='ALERTA!',text='Boleto salvo em dowloads')

    def mensalidade_do_mes(self):
        for indice in range(0, len(self.valor_mensalidade)):
            linha = [self.valor_mensalidade[indice].text]
            self.planilha_dados.append(linha)
        self.planilha.save('Mensalidade.xlsx')

    def organizarmensalidades(self):
        os.mkdir('/home/duduka/Área de Trabalho/Boletos')
        os.chdir('/home/duduka/Área de Trabalho/Boletos')
        shutil.move('/home/duduka/Downloads/Boleto.pdf', '/home/duduka/Área de Trabalho/Boletos')



    def armazernar_dados(self):
        pass


    def acessar(self):
        self.driver.get('https://login.uniasselvi.com.br/index.php')


if __name__ == '__main__':
    start = Extracao()
    start.acessar()
    start.scraping()
